#!/usr/bin/env python3
"""
Sales Register Converter
========================
Converts the raw accounting software "GSTR1_SYSTEM_DATA" file into the
clean single-sheet portal / sales-register format that you previously edited
by hand each month.

What it does
------------
• Reads every group section (registered, unregistered, tax-free, …) exactly
  as the accounting software exported them.
• Removes the combined sub-total rows (the extra "TOTAL FOR GST XX % :" line
  that sums two sections together).
• Moves the individual section total labels from column B → column C and
  derives them cleanly from the section heading (e.g. "GST - GST 18 %" →
  "TOTAL FOR GST 18 %").
• Renames two column headers:  "Bill No" → "Bill.No"  and  "Bill Amt" → "Bill AMt".
• Writes a clean GRAND TOTAL at the end.

Usage
-----
    python3 sales_register_converter.py  <system_data.xlsx>  [output.xlsx]

    If output.xlsx is omitted the output file is saved as
    <COMPANY_CODE>_PORTAL.xlsx  (e.g. BDB_PORTAL.xlsx).

Examples
--------
    python3 sales_register_converter.py  BDB_GSTR1_SYSTEM_DATA_APRIL-2026.xlsx
    python3 sales_register_converter.py  BDB_GSTR1_SYSTEM_DATA_APRIL-2026.xlsx  BDB_APRIL-2026.xlsx

    # Process all 4 companies at once (bash):
    for co in BDB DPB HKI HTEI; do
        python3 sales_register_converter.py  ${co}_GSTR1_SYSTEM_DATA_APRIL-2026.xlsx  ${co}.xlsx
    done
"""

import sys
import os
import re
from openpyxl import load_workbook, Workbook


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def to_f(v):
    """Safe float; returns 0.0 on failure."""
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def is_empty_row(row):
    return not any(str(c or '').strip() for c in row)


def is_group_header_row(row):
    """
    A group header row has:
      col[0] empty, col[1] empty, col[2] non-empty text (the group name),
      col[4] empty (not an invoice amount).
    """
    c0 = str(row[0] or '').strip()
    c1 = str(row[1] or '').strip()
    c2 = str(row[2] or '').strip()
    c4 = str(row[4] or '').strip()
    return (not c0) and (not c1) and bool(c2) and (not c4)


def is_grand_total_row(row):
    return 'GRAND TOTAL' in str(row[1] or '').upper()


def is_total_row(row):
    """Any sub-total (section total or combined total)."""
    c1 = str(row[1] or '').strip().upper()
    return c1.startswith('TOTAL') or 'GRAND TOTAL' in c1


def is_unreg_section(header):
    h = header.upper()
    return 'UN REG' in h or 'UNREG' in h or 'UN REGI' in h


def reorder_sections(sections):
    """
    Move any section whose header starts with 'TXF' to just before the first
    UNREG section that currently appears *before* it.

    This matches the manual convention used for companies like BDB where the
    accounting software emits:  GST-REG → UNREG → TXF-TAXFREE
    but the portal file needs:  GST-REG → TXF-TAXFREE → UNREG
    Companies that use the 'TF' prefix (not 'TXF') are left in their original
    order.
    """
    result = list(sections)

    # Process in reverse so indices remain valid after each insertion/removal
    txf_indices = [i for i, (h, _) in enumerate(result) if h.upper().startswith('TXF')]

    for txf_idx in reversed(txf_indices):
        # Is there any UNREG section sitting before this TXF?
        preceding_unreg = [j for j in range(txf_idx)
                           if is_unreg_section(result[j][0])]
        if not preceding_unreg:
            continue   # nothing to reorder
        first_unreg = min(preceding_unreg)
        txf_sec = result.pop(txf_idx)
        result.insert(first_unreg, txf_sec)

    return result


def derive_total_label(group_header):
    """
    "GST - GST 18 %"        →  "TOTAL FOR GST 18 %"
    "GS1 - GST 18% UN REG." →  "TOTAL FOR GST 18% UN REG."
    "TF - TAX FREE"         →  "TOTAL FOR TAX FREE"
    "TXF - TAX FREE"        →  "TOTAL FOR TAX FREE"
    """
    gh = str(group_header).strip()
    if ' - ' in gh:
        suffix = gh.split(' - ', 1)[1].strip()
    elif '-' in gh:
        suffix = gh.split('-', 1)[1].strip()
    else:
        suffix = gh
    return f"TOTAL FOR {suffix}"


# ──────────────────────────────────────────────────────────────────────────────
# STEP 1 – PARSE SYSTEM DATA FILE
# ──────────────────────────────────────────────────────────────────────────────

def parse_system_data(filepath):
    """
    Returns
    -------
    company_info : list[list]   rows 0-2 (company name, "Sales Register Printing", period)
    col_headers  : list         original column-header row from the file
    sections     : list of (group_header_str, [invoice_rows])
                   invoice_rows are the 13-column data rows, copied as-is
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # ── locate column-header row (has "Gst No" / "Party Name" / "Bill No")
    header_row_idx = None
    for i, row in enumerate(rows):
        cells = [str(c or '').strip() for c in row]
        if any(k in cells for k in ('Gst No', 'Party Name', 'Bill No')):
            header_row_idx = i
            break
    if header_row_idx is None:
        header_row_idx = 3   # fallback

    company_info = rows[:3]              # rows 0, 1, 2 (always 3 rows)
    col_headers  = list(rows[header_row_idx])

    # ── parse sections ────────────────────────────────────────────────────────
    sections        = []
    current_group   = None
    current_invoices= []
    last_was_total  = False   # True right after we've seen a section-total row

    for row in rows[header_row_idx + 1:]:
        # Ensure at least 13 columns
        while len(row) < 13:
            row.append(None)

        if is_empty_row(row):
            continue                          # skip blank rows

        if is_grand_total_row(row):
            # Close any open section, then stop
            if current_group is not None:
                sections.append((current_group, current_invoices))
            break

        if is_total_row(row):
            if last_was_total:
                # ── combined sub-total: a TOTAL that immediately follows
                #    another TOTAL (with at most empty rows in between).
                #    These combine two sections (e.g. REG + UNREG).
                #    We skip them entirely.
                pass                          # do NOT update last_was_total
            else:
                # ── genuine section total: close the current section
                if current_group is not None:
                    sections.append((current_group, current_invoices))
                    current_group    = None
                    current_invoices = []
                last_was_total = True
            continue

        if is_group_header_row(row):
            # Start a new section
            last_was_total = False
            current_group  = str(row[2] or '').strip()
            current_invoices = []
            continue

        # ── invoice row ───────────────────────────────────────────────────────
        last_was_total = False
        if current_group is not None:
            current_invoices.append([v for v in row[:13]])

    # Edge-case: last section had no closing total (shouldn't happen but guard)
    if current_group is not None and current_invoices:
        sections.append((current_group, current_invoices))

    return company_info, col_headers, sections


# ──────────────────────────────────────────────────────────────────────────────
# STEP 2 – WRITE SIMPLE PORTAL FILE
# ──────────────────────────────────────────────────────────────────────────────

def write_simple_file(company_info, col_headers, sections, output_path):
    """
    Output layout (0-indexed rows → Excel rows):
      R0  → row 1  : company name
      R1  → row 2  : "Sales Register Printing"
      R2  → row 3  : date range
      R3  → row 4  : (empty)
      R4  → row 5  : column headers  (Bill.No / Bill AMt)
      R5+ → row 6+ : sections  →  group header | invoices | total | (empty)
      last         : GRAND TOTAL
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # ── company header rows (rows 1–3) ────────────────────────────────────────
    for r_idx, crow in enumerate(company_info[:3]):
        for c_idx, val in enumerate(crow[:13]):
            ws.cell(row=r_idx + 1, column=c_idx + 1).value = val

    # row 4 is intentionally left empty

    # ── column header row (row 5) ─────────────────────────────────────────────
    hdrs = list(col_headers)
    while len(hdrs) < 13:
        hdrs.append(None)
    hdrs[2] = 'Bill.No'    # was "Bill No"
    hdrs[4] = 'Bill AMt'   # was "Bill Amt"
    for c_idx, val in enumerate(hdrs[:13]):
        ws.cell(row=5, column=c_idx + 1).value = val

    # ── sections (starting at Excel row 6) ───────────────────────────────────
    excel_row = 6
    grand = dict(bill=0.0, sgst=0.0, cgst=0.0, igst=0.0, taxable=0.0)

    for group_hdr, invoices in reorder_sections(sections):
        # Group header (col C = column 3)
        ws.cell(excel_row, 3).value = group_hdr
        excel_row += 1

        # Invoice rows (copy all 13 columns exactly as in source)
        for inv in invoices:
            for c_idx, val in enumerate(inv[:13]):
                ws.cell(excel_row, c_idx + 1).value = val
            excel_row += 1

        # Section sums (computed from invoices, not copied from source totals)
        bill    = sum(to_f(i[4])  for i in invoices)
        sgst    = sum(to_f(i[7])  for i in invoices)
        cgst    = sum(to_f(i[9])  for i in invoices)
        igst    = sum(to_f(i[11]) for i in invoices)
        taxable = sum(to_f(i[12]) for i in invoices)

        # Total row: label in col C, amounts in the same columns as invoices.
        # Always write numeric 0 for zero amounts — never leave them blank.
        ws.cell(excel_row, 3).value  = derive_total_label(group_hdr)
        ws.cell(excel_row, 5).value  = bill
        ws.cell(excel_row, 8).value  = sgst
        ws.cell(excel_row, 10).value = cgst
        ws.cell(excel_row, 12).value = igst
        ws.cell(excel_row, 13).value = taxable
        excel_row += 1

        # One empty separator row
        excel_row += 1

        grand['bill']    += bill
        grand['sgst']    += sgst
        grand['cgst']    += cgst
        grand['igst']    += igst
        grand['taxable'] += taxable

    # ── GRAND TOTAL row ───────────────────────────────────────────────────────
    ws.cell(excel_row, 3).value  = 'GRAND TOTAL'
    ws.cell(excel_row, 5).value  = grand['bill']
    ws.cell(excel_row, 8).value  = grand['sgst']
    ws.cell(excel_row, 10).value = grand['cgst']
    ws.cell(excel_row, 12).value = grand['igst']
    ws.cell(excel_row, 13).value = grand['taxable']

    wb.save(output_path)
    return grand


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def derive_output_name(input_path):
    """BDB_GSTR1_SYSTEM_DATA_APRIL-2026.xlsx  →  BDB_PORTAL.xlsx"""
    base = os.path.basename(input_path)
    name = os.path.splitext(base)[0]
    # Extract company code (part before first underscore)
    company_code = name.split('_')[0]
    return f"{company_code}_PORTAL.xlsx"


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("ERROR: please provide the system-data file path.")
        sys.exit(1)

    input_file  = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else derive_output_name(input_file)

    if not os.path.exists(input_file):
        print(f"ERROR: File not found: {input_file}")
        sys.exit(1)

    print(f"Reading  : {input_file}")
    company_info, col_headers, sections = parse_system_data(input_file)

    company_name = str(company_info[0][0] if company_info else '').strip()
    period       = str(company_info[2][0] if len(company_info) > 2 else '').strip()
    print(f"Company  : {company_name}")
    print(f"Period   : {period}")
    print(f"Sections : {len(sections)}")
    for hdr, invs in sections:
        print(f"           {hdr!r:45s} → {len(invs)} invoices")

    grand = write_simple_file(company_info, col_headers, sections, output_file)

    print(f"\nSaved    : {output_file}")
    print(f"Grand totals → Bill: {grand['bill']:,.2f} | "
          f"SGST: {grand['sgst']:,.2f} | CGST: {grand['cgst']:,.2f} | "
          f"IGST: {grand['igst']:,.2f} | Taxable: {grand['taxable']:,.2f}")


if __name__ == '__main__':
    main()
