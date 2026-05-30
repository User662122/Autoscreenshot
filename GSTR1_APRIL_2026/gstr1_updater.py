#!/usr/bin/env python3
"""
GSTR-1 Updater Script
---------------------
Reads raw accounting software GSTR-1 export and:
  1. Updates the existing GSTR-1 template (replaces only data rows, keeps all formatting)
  2. Generates a reformatted single-sheet portal/sales-register file

Usage:
    python gstr1_updater.py <raw_file.xlsx> <gstr1_template.xlsx> [output_prefix]

    output_prefix defaults to "GSTR1_OUTPUT"

Outputs:
    <output_prefix>_GSTR1.xlsx   - Updated GSTR-1 multi-sheet template
    <output_prefix>_PORTAL.xlsx  - Reformatted sales register (REG -> TAX FREE -> UNREG order)
"""

import sys
import os
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def to_f(v):
    """Safe float conversion."""
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def normalize_date(val):
    """Return date value suitable for openpyxl (pass-through integer or datetime)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        return int(val)   # keep as Excel serial number
    return val


# ──────────────────────────────────────────────────────────────────────────────
# STEP 1: PARSE RAW ACCOUNTING SOFTWARE FILE
# ──────────────────────────────────────────────────────────────────────────────

GROUP_KEYWORDS = ['GST', 'TXF', 'TAX FREE', 'GS1', 'GUR', 'G18', 'UNR']

def is_group_header(row):
    """True if this row is a section/group header (not an invoice)."""
    gst_no = str(row[0] or '').strip()
    party  = str(row[1] or '').strip()
    bill   = str(row[2] or '').strip()
    if gst_no or party:
        return False
    if not bill:
        return False
    # Must not be a pure number
    if bill.replace('.', '').isnumeric():
        return False
    # Must match one of the group keywords
    bill_up = bill.upper()
    return any(k in bill_up for k in GROUP_KEYWORDS)


def is_total_row(row):
    """True if this row is a total/summary row."""
    party  = str(row[1] or '').upper()
    bill   = str(row[2] or '').upper()
    return 'TOTAL' in party or 'TOTAL' in bill or 'GRAND TOTAL' in party


def parse_raw_file(filepath):
    """
    Parse accounting software GSTR-1 export file.

    Returns dict:
        company  – company name string
        period   – period string
        b2b      – list of registered (B2B) invoices
        b2cs     – list of unregistered taxable (B2CS) invoices
        exempt   – list of nil/tax-free (Exempt) invoices
        sections – list of (group_header_str, [invoices]) in raw file order
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    company = str(rows[0][0] or '').strip()
    period  = str(rows[2][0] or '').strip() if len(rows) > 2 else ''

    # Find column-header row ("Gst No", "Party Name", …)
    header_idx = 3
    for i, row in enumerate(rows):
        cells = [str(c or '').strip() for c in row]
        if 'Party Name' in cells or 'Gst No' in cells:
            header_idx = i
            break

    b2b    = []
    b2cs   = []
    exempt = []

    sections        = []
    current_group   = ''
    current_section = []

    for row in rows[header_idx + 1:]:
        # Pad short rows
        while len(row) < 13:
            row.append(None)

        if is_group_header(row):
            # Save previous section
            if current_section:
                sections.append((current_group, current_section))
                current_section = []
            current_group = str(row[2] or '').strip()
            continue

        if is_total_row(row):
            continue

        gst_no   = str(row[0] or '').strip()
        party    = str(row[1] or '').strip()
        bill_no  = row[2]
        date_val = row[3]
        bill_amt = row[4]
        state    = str(row[5] or '').strip() or '24-Gujarat'
        sgst_pct = row[6];  sgst_amt = row[7]
        cgst_pct = row[8];  cgst_amt = row[9]
        igst_pct = row[10]; igst_amt = row[11]
        taxable  = row[12]

        if not party or bill_no is None:
            continue

        sgst_p = to_f(sgst_pct); cgst_p = to_f(cgst_pct); igst_p = to_f(igst_pct)
        sgst_a = to_f(sgst_amt); cgst_a = to_f(cgst_amt); igst_a = to_f(igst_amt)
        bill_a = to_f(bill_amt); tax_a  = to_f(taxable)

        rate = (sgst_p + cgst_p) if (sgst_p + cgst_p) > 0 else igst_p

        inv = {
            'gst_no'   : gst_no,
            'party'    : party,
            'bill_no'  : bill_no,
            'date'     : normalize_date(date_val),
            'bill_amt' : bill_a,
            'state'    : state,
            'rate'     : rate,
            'sgst_pct' : sgst_p, 'sgst_amt': sgst_a,
            'cgst_pct' : cgst_p, 'cgst_amt': cgst_a,
            'igst_pct' : igst_p, 'igst_amt': igst_a,
            'taxable'  : tax_a,
            'group'    : current_group,
        }

        current_section.append(inv)

        # Classify
        if len(gst_no) == 15:
            b2b.append(inv)
        elif rate == 0:
            exempt.append(inv)
        else:
            b2cs.append(inv)

    if current_section:
        sections.append((current_group, current_section))

    return {
        'company' : company,
        'period'  : period,
        'b2b'     : b2b,
        'b2cs'    : b2cs,
        'exempt'  : exempt,
        'sections': sections,
    }


# ──────────────────────────────────────────────────────────────────────────────
# STEP 2: UPDATE GSTR-1 TEMPLATE SHEETS
# ──────────────────────────────────────────────────────────────────────────────

def overwrite_data_rows(ws, new_rows, data_start_row=5):
    """
    Replace data in sheet starting at data_start_row.
    Overwrites rows that have new data; clears leftover old rows.
    Does NOT delete rows (preserves any named ranges / merges outside data area).
    """
    # Find actual last row with data (scan limited rows, not 20k empties)
    old_last = data_start_row - 1
    for r in range(data_start_row, data_start_row + 300):
        if any(ws.cell(r, c).value is not None for c in range(1, 15)):
            old_last = r

    # Write new data rows
    for i, row_data in enumerate(new_rows):
        r = data_start_row + i
        for j, val in enumerate(row_data):
            ws.cell(row=r, column=j + 1).value = val
        for j in range(len(row_data), 14):
            ws.cell(row=r, column=j + 1).value = None

    # Delete leftover old data rows
    new_last = data_start_row + len(new_rows)
    if old_last >= new_last:
        ws.delete_rows(new_last, old_last - new_last + 1)

    # Trim pre-allocated empty rows so the file saves fast
    trim_from = data_start_row + len(new_rows) + 5
    if ws.max_row > trim_from + 20:
        try:
            ws.delete_rows(trim_from, ws.max_row - trim_from + 1)
        except Exception:
            pass


# ── b2b,sez,de ────────────────────────────────────────────────────────────────

def update_b2b(ws, b2b):
    DATA_START   = 5
    SUMMARY_ROW  = 3

    rows = []
    for inv in b2b:
        rows.append([
            inv['gst_no'],
            inv['party'],
            inv['bill_no'],
            inv['date'],
            inv['bill_amt'],
            inv['state'],
            'N',              # Reverse Charge
            '',               # Applicable % of Tax Rate
            'Regular B2B',   # Invoice Type
            '',               # E-Commerce GSTIN
            inv['rate'],
            inv['taxable'],
            '',               # Cess
            '',
        ])

    overwrite_data_rows(ws, rows, DATA_START)

    unique_gstins = len(set(i['gst_no'] for i in b2b))
    total_inv_val = sum(i['bill_amt']  for i in b2b)
    total_taxable = sum(i['taxable']   for i in b2b)

    ws.cell(SUMMARY_ROW, 1).value  = unique_gstins
    ws.cell(SUMMARY_ROW, 3).value  = len(b2b)
    ws.cell(SUMMARY_ROW, 5).value  = total_inv_val
    ws.cell(SUMMARY_ROW, 12).value = total_taxable
    ws.cell(SUMMARY_ROW, 13).value = 0
    print(f"    b2b,sez,de : {len(b2b)} invoices | {unique_gstins} recipients | taxable {total_taxable:,.0f}")


# ── b2cs ──────────────────────────────────────────────────────────────────────

def update_b2cs(ws, b2cs):
    DATA_START  = 5
    SUMMARY_ROW = 3

    # Aggregate by state + rate
    groups = {}
    for inv in b2cs:
        key = (inv['state'], inv['rate'])
        groups[key] = groups.get(key, 0.0) + inv['taxable']

    rows = []
    for (state, rate), taxable in sorted(groups.items()):
        rows.append(['OE', state, '', rate, taxable, '', '', ''])

    overwrite_data_rows(ws, rows, DATA_START)

    total_taxable = sum(i['taxable'] for i in b2cs)
    ws.cell(SUMMARY_ROW, 5).value = total_taxable
    ws.cell(SUMMARY_ROW, 6).value = 0
    print(f"    b2cs       : {len(b2cs)} invoices | {len(groups)} groups | taxable {total_taxable:,.0f}")


# ── exemp ─────────────────────────────────────────────────────────────────────

def update_exemp(ws, exempt):
    DATA_START  = 5
    SUMMARY_ROW = 3

    # Read existing description labels from template to preserve them
    existing_descs = []
    for r in range(DATA_START, ws.max_row + 1):
        desc = ws.cell(r, 1).value
        if desc and str(desc).strip():
            existing_descs.append(str(desc).strip())

    # Group exempt invoices by bill-number series prefix
    series_totals = {}
    for inv in exempt:
        bill = str(inv['bill_no']).strip()
        prefix = bill.split('/')[0].upper() if '/' in bill else re.sub(r'\d', '', bill).upper() or 'MISC'
        series_totals[prefix] = series_totals.get(prefix, 0.0) + inv['taxable']

    # Map series → description using template labels
    DESC_KEYWORDS = {
        'RES'  : 'RESIDENTIAL',
        'WIM'  : 'WIND',
        'MISC' : '',
    }
    used_prefixes = set()

    rows = []
    if existing_descs:
        for desc in existing_descs:
            du = desc.upper()
            assigned_amt = 0.0
            for prefix, amt in series_totals.items():
                if prefix in used_prefixes:
                    continue
                kw = DESC_KEYWORDS.get(prefix, prefix)
                if kw and kw in du:
                    assigned_amt = amt
                    used_prefixes.add(prefix)
                    break
            # Fall back: if only one series and no match yet, use it for first desc
            if assigned_amt == 0 and not used_prefixes and series_totals:
                prefix, amt = next(iter(series_totals.items()))
                assigned_amt = amt
                used_prefixes.add(prefix)

            rows.append([desc, '', assigned_amt if assigned_amt else '', ''])

        # Add any leftover series not mapped to existing descriptions
        for prefix, amt in series_totals.items():
            if prefix not in used_prefixes:
                rows.append([f'{prefix} INCOME', '', amt, ''])
    else:
        # No existing descriptions — generate from series
        label_map = {'RES': 'RESIDENTIAL RENT', 'WIM': 'WIND MILL INCOME'}
        for prefix, amt in series_totals.items():
            rows.append([label_map.get(prefix, f'{prefix} INCOME'), '', amt, ''])

    overwrite_data_rows(ws, rows, DATA_START)

    total_exempt = sum(i['taxable'] for i in exempt)
    ws.cell(SUMMARY_ROW, 2).value = 0
    ws.cell(SUMMARY_ROW, 3).value = total_exempt
    ws.cell(SUMMARY_ROW, 4).value = 0
    print(f"    exemp      : {len(exempt)} invoices | taxable {total_exempt:,.0f}")


# ── hsn(b2b) ──────────────────────────────────────────────────────────────────

def update_hsn_b2b(ws, b2b, exempt):
    DATA_START  = 5
    SUMMARY_ROW = 3

    # Read existing HSN rows from template
    existing = []
    for r in range(DATA_START, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 12)]
        if any(v is not None and str(v).strip() for v in vals):
            existing.append(vals)

    # Compute totals
    b2b_qty   = len(b2b)
    b2b_val   = sum(i['bill_amt']  for i in b2b)
    b2b_tax   = sum(i['taxable']   for i in b2b)
    b2b_igst  = sum(i['igst_amt']  for i in b2b)
    b2b_cgst  = sum(i['cgst_amt']  for i in b2b)
    b2b_sgst  = sum(i['sgst_amt']  for i in b2b)
    b2b_rate  = b2b[0]['rate'] if b2b else 18

    ex_qty    = len(exempt)
    ex_val    = sum(i['bill_amt']  for i in exempt)
    ex_tax    = sum(i['taxable']   for i in exempt)

    b2b_done  = False
    ex_done   = False
    new_rows  = []

    for rv in existing:
        hsn  = rv[0]
        desc = str(rv[1] or '').strip()
        uqc  = rv[2]
        du   = desc.upper()

        is_empty = not hsn and not desc

        if is_empty:
            new_rows.append([None, None, None, '', '', '', '', '', '', '', ''])

        elif 'RESIDENTIAL' in du or ('WIND' in du and b2b_done):
            # Exempt / residential / wind-mill row
            if not ex_done and exempt:
                new_rows.append([hsn, desc, uqc, ex_qty, ex_val, 0, ex_tax, 0, 0, 0, 0])
                ex_done = True
            else:
                new_rows.append([hsn, desc, uqc, '', '', '', '', '', '', '', ''])

        else:
            # Main taxable (B2B) row
            if not b2b_done and b2b:
                new_rows.append([hsn, desc, uqc, b2b_qty, b2b_val, b2b_rate, b2b_tax,
                                 b2b_igst, b2b_cgst, b2b_sgst, ''])
                b2b_done = True
            else:
                new_rows.append([hsn, desc, uqc, '', '', '', '', '', '', '', ''])

    # No existing rows — create default
    if not new_rows:
        if b2b:
            new_rows.append([997212, 'RENT INCOME', '', b2b_qty, b2b_val, b2b_rate,
                             b2b_tax, b2b_igst, b2b_cgst, b2b_sgst, ''])
        if exempt:
            new_rows.append([997212, 'RENT INCOME RESIDENTIAL', '', ex_qty, ex_val,
                             0, ex_tax, 0, 0, 0, 0])

    overwrite_data_rows(ws, new_rows, DATA_START)

    unique_hsns   = len(set(str(r[0]) for r in new_rows if r[0]))
    total_val     = b2b_val + ex_val
    total_tax     = b2b_tax + ex_tax
    ws.cell(SUMMARY_ROW, 1).value  = unique_hsns
    ws.cell(SUMMARY_ROW, 5).value  = total_val
    ws.cell(SUMMARY_ROW, 7).value  = total_tax
    ws.cell(SUMMARY_ROW, 8).value  = b2b_igst
    ws.cell(SUMMARY_ROW, 9).value  = b2b_cgst
    ws.cell(SUMMARY_ROW, 10).value = b2b_sgst
    ws.cell(SUMMARY_ROW, 11).value = 0
    print(f"    hsn(b2b)   : {b2b_qty} b2b + {ex_qty} exempt rows")


# ── hsn(b2c) ──────────────────────────────────────────────────────────────────

def update_hsn_b2c(ws, b2cs):
    DATA_START  = 5
    SUMMARY_ROW = 3

    # Read first existing HSN row (to preserve HSN code + description)
    existing_row = None
    for r in range(DATA_START, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 12)]
        if any(v is not None and str(v).strip() for v in vals):
            existing_row = vals
            break

    qty  = len(b2cs)
    val  = sum(i['bill_amt']  for i in b2cs)
    tax  = sum(i['taxable']   for i in b2cs)
    igst = sum(i['igst_amt']  for i in b2cs)
    cgst = sum(i['cgst_amt']  for i in b2cs)
    sgst = sum(i['sgst_amt']  for i in b2cs)
    rate = b2cs[0]['rate'] if b2cs else 18

    if existing_row:
        hsn  = existing_row[0]
        desc = existing_row[1]
        uqc  = existing_row[2]
    else:
        hsn, desc, uqc = 997212, 'RENT INCOME', ''

    new_rows = []
    if b2cs:
        new_rows.append([hsn, desc, uqc, qty, val, rate, tax, igst, cgst, sgst, ''])

    overwrite_data_rows(ws, new_rows, DATA_START)

    ws.cell(SUMMARY_ROW, 1).value  = 1 if new_rows else 0
    ws.cell(SUMMARY_ROW, 5).value  = val
    ws.cell(SUMMARY_ROW, 7).value  = tax
    ws.cell(SUMMARY_ROW, 8).value  = igst
    ws.cell(SUMMARY_ROW, 9).value  = cgst
    ws.cell(SUMMARY_ROW, 10).value = sgst
    ws.cell(SUMMARY_ROW, 11).value = 0
    print(f"    hsn(b2c)   : {qty} invoices | taxable {tax:,.0f}")


# ── docs ──────────────────────────────────────────────────────────────────────

def bill_series_key(bill_no):
    """Return series prefix key for a bill number."""
    b = str(bill_no).strip()
    if '/' in b:
        return b.split('/')[0].upper() + '/'
    m = re.match(r'^[A-Za-z-]+', b)
    return m.group(0).upper() if m else 'NUMERIC'


def update_docs(ws, b2b, b2cs, exempt):
    DATA_START  = 5
    SUMMARY_ROW = 3

    all_inv = b2b + b2cs + exempt

    # Group by series
    series = {}
    for inv in all_inv:
        key = bill_series_key(inv['bill_no'])
        if key not in series:
            series[key] = []
        series[key].append(inv['bill_no'])

    rows = []
    total_docs = 0
    for key in sorted(series.keys()):
        bills = series[key]
        total_docs += len(bills)

        if key == 'NUMERIC':
            nums = sorted(int(float(str(b))) for b in bills)
            from_v, to_v = nums[0], nums[-1]
        else:
            sorted_b = sorted(str(b) for b in bills)
            from_v, to_v = sorted_b[0], sorted_b[-1]

        rows.append(['Invoices for outward supply', from_v, to_v, len(bills), ''])

    overwrite_data_rows(ws, rows, DATA_START)

    ws.cell(SUMMARY_ROW, 4).value = total_docs
    ws.cell(SUMMARY_ROW, 5).value = 0
    print(f"    docs       : {total_docs} docs in {len(rows)} series")


# ── Main update orchestrator ──────────────────────────────────────────────────

def update_gstr1_template(raw_data, template_path, output_path):
    """Open template, replace data in each sheet, save."""
    print(f"\n  Loading template: {template_path}")
    wb = load_workbook(template_path)

    b2b    = raw_data['b2b']
    b2cs   = raw_data['b2cs']
    exempt = raw_data['exempt']
    print(f"  Invoices: B2B={len(b2b)}  B2CS={len(b2cs)}  Exempt={len(exempt)}")

    SHEET_UPDATERS = {
        'b2b,sez,de': lambda ws: update_b2b(ws, b2b),
        'b2cs'       : lambda ws: update_b2cs(ws, b2cs),
        'exemp'      : lambda ws: update_exemp(ws, exempt),
        'hsn(b2b)'   : lambda ws: update_hsn_b2b(ws, b2b, exempt),
        'hsn(b2c)'   : lambda ws: update_hsn_b2c(ws, b2cs),
        'docs'       : lambda ws: update_docs(ws, b2b, b2cs, exempt),
    }

    for sheet_name, updater in SHEET_UPDATERS.items():
        if sheet_name in wb.sheetnames:
            updater(wb[sheet_name])
        else:
            print(f"    (skipping missing sheet: {sheet_name})")

    wb.save(output_path)
    print(f"  Saved → {output_path}")


# ──────────────────────────────────────────────────────────────────────────────
# STEP 3: GENERATE PORTAL / SALES-REGISTER FILE
# ──────────────────────────────────────────────────────────────────────────────

def section_type(group_header):
    """Classify a section as 'reg', 'unreg', or 'taxfree'."""
    g = group_header.upper()
    if 'TAX FREE' in g or 'TXF' in g or re.match(r'^TF\b', g):
        return 'taxfree'
    if 'UN REG' in g or 'UNREG' in g or 'GS1' in g or 'GUR' in g:
        return 'unreg'
    return 'reg'


def portal_total_label(group_header):
    """Generate portal-style total label from group header."""
    g = group_header
    # Extract part after dash if present
    if ' - ' in g:
        label = g.split(' - ', 1)[1].strip()
    elif '-' in g:
        label = g.split('-', 1)[1].strip()
    else:
        label = g.strip()
    # Remove trailing colon/period/spaces then build label
    label = re.sub(r'[\s:]+$', '', label)
    return f"TOTAL FOR {label}"


def generate_portal_file(raw_data, output_path):
    """Generate reorganised single-sheet sales-register portal file."""
    sections = raw_data['sections']

    # Classify and reorder: REG → TAX FREE → UNREG
    regs     = [(h, i) for h, i in sections if section_type(h) == 'reg']
    taxfrees = [(h, i) for h, i in sections if section_type(h) == 'taxfree']
    unregs   = [(h, i) for h, i in sections if section_type(h) == 'unreg']
    ordered  = regs + taxfrees + unregs

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Company header (3 rows)
    ws.cell(1, 1).value = raw_data['company']
    ws.cell(2, 1).value = 'Sales Register Printing'
    ws.cell(3, 1).value = raw_data['period']
    ws.cell(4, 1).value = ''

    # Column headers row (row 5 as blank then row 5 is header)
    headers = ['Gst No', 'Party Name', 'Bill.No', 'Date', 'Bill AMt', 'State',
               'SGST %', 'SGST Amt', 'CGST %', 'CGST Amt',
               'IGST %', 'IGST Amt', 'Taxable Amt']
    for j, h in enumerate(headers):
        ws.cell(5, j + 1).value = h

    cur_row = 6
    g_bill = g_sgst = g_cgst = g_igst = g_tax = 0.0

    for group_hdr, invoices in ordered:
        # Empty separator
        ws.cell(cur_row, 1).value = ''
        cur_row += 1

        # Group header row
        ws.cell(cur_row, 3).value = group_hdr
        cur_row += 1

        sec_bill = sec_sgst = sec_cgst = sec_igst = sec_tax = 0.0

        for inv in invoices:
            ws.cell(cur_row, 1).value  = inv['gst_no']
            ws.cell(cur_row, 2).value  = inv['party']
            ws.cell(cur_row, 3).value  = inv['bill_no']
            ws.cell(cur_row, 4).value  = inv['date']
            ws.cell(cur_row, 5).value  = inv['bill_amt']
            ws.cell(cur_row, 6).value  = inv['state']
            ws.cell(cur_row, 7).value  = inv['sgst_pct']
            ws.cell(cur_row, 8).value  = inv['sgst_amt']
            ws.cell(cur_row, 9).value  = inv['cgst_pct']
            ws.cell(cur_row, 10).value = inv['cgst_amt']
            ws.cell(cur_row, 11).value = inv['igst_pct']
            ws.cell(cur_row, 12).value = inv['igst_amt']
            ws.cell(cur_row, 13).value = inv['taxable']
            cur_row += 1

            sec_bill += inv['bill_amt'];  sec_sgst += inv['sgst_amt']
            sec_cgst += inv['cgst_amt'];  sec_igst += inv['igst_amt']
            sec_tax  += inv['taxable']

        # Section total row
        total_lbl = portal_total_label(group_hdr)
        ws.cell(cur_row, 3).value  = total_lbl
        ws.cell(cur_row, 5).value  = sec_bill
        ws.cell(cur_row, 8).value  = sec_sgst
        ws.cell(cur_row, 10).value = sec_cgst
        ws.cell(cur_row, 12).value = sec_igst
        ws.cell(cur_row, 13).value = sec_tax
        cur_row += 1

        g_bill += sec_bill;  g_sgst += sec_sgst
        g_cgst += sec_cgst;  g_igst += sec_igst
        g_tax  += sec_tax

    # Empty row before grand total
    cur_row += 1

    # Grand total
    ws.cell(cur_row, 3).value  = 'GRAND TOTAL'
    ws.cell(cur_row, 5).value  = g_bill
    ws.cell(cur_row, 8).value  = g_sgst
    ws.cell(cur_row, 10).value = g_cgst
    ws.cell(cur_row, 12).value = g_igst
    ws.cell(cur_row, 13).value = g_tax

    wb.save(output_path)
    print(f"  Saved → {output_path}")


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print(__doc__)
        print("ERROR: provide <raw_file.xlsx> <gstr1_template.xlsx> [output_prefix]")
        sys.exit(1)

    raw_file      = sys.argv[1]
    template_file = sys.argv[2]
    prefix        = sys.argv[3] if len(sys.argv) > 3 else 'GSTR1_OUTPUT'

    for f in (raw_file, template_file):
        if not os.path.exists(f):
            print(f"ERROR: File not found: {f}")
            sys.exit(1)

    gstr1_out  = f"{prefix}_GSTR1.xlsx"
    portal_out = f"{prefix}_PORTAL.xlsx"

    print("=" * 60)
    print(f"GSTR-1 Updater")
    print("=" * 60)
    print(f"  Raw file  : {raw_file}")
    print(f"  Template  : {template_file}")

    raw_data = parse_raw_file(raw_file)
    print(f"  Company   : {raw_data['company']}")
    print(f"  Period    : {raw_data['period']}")

    print("\n[1/2] Updating GSTR-1 template...")
    update_gstr1_template(raw_data, template_file, gstr1_out)

    print("\n[2/2] Generating portal/sales-register file...")
    generate_portal_file(raw_data, portal_out)

    print("\n" + "=" * 60)
    print("DONE")
    print(f"  GSTR-1 template : {gstr1_out}")
    print(f"  Portal file     : {portal_out}")
    print("=" * 60)


if __name__ == '__main__':
    main()
